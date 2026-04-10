"""
Excel Exporter - Export Analytics & Feedback to Excel
Copyright (c) 2026 Sargam Chicholikar
"""

import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    def __init__(self):
        self.exports_dir = "exports"
        os.makedirs(self.exports_dir, exist_ok=True)
        print("✅ Excel Exporter ready")
    
    def export_analytics(self):
        """Export analytics to Excel"""
        
        try:
            # Load analytics data
            with open('analytics/usage_stats.json', 'r') as f:
                analytics = json.load(f)
            
            # Create workbook
            wb = Workbook()
            
            # Sheet 1: Overview
            ws1 = wb.active
            ws1.title = "Overview"
            
            # Header styling
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            # Overview Stats
            ws1['A1'] = 'MediSimplify - Analytics Report'
            ws1['A1'].font = Font(bold=True, size=16)
            ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            
            ws1['A4'] = 'Metric'
            ws1['B4'] = 'Value'
            ws1['A4'].fill = header_fill
            ws1['B4'].fill = header_fill
            ws1['A4'].font = header_font
            ws1['B4'].font = header_font
            
            # Data
            metrics = [
                ('Total Analyses', analytics.get('total_analyses', 0)),
                ('Prescriptions', analytics.get('prescriptions_analyzed', 0)),
                ('Lab Reports', analytics.get('labs_analyzed', 0)),
                ('X-Rays', analytics.get('xrays_analyzed', 0)),
                ('Manual Drug Entries', analytics.get('manual_drug_entries', 0)),
                ('Voice Playbacks', analytics['feature_usage'].get('voice_playback', 0)),
                ('Multi-Report Analyses', analytics['feature_usage'].get('multi_report_analysis', 0))
            ]
            
            row = 5
            for metric, value in metrics:
                ws1[f'A{row}'] = metric
                ws1[f'B{row}'] = value
                ws1[f'B{row}'].font = Font(bold=True, size=12)
                row += 1
            
            # Auto-width
            ws1.column_dimensions['A'].width = 30
            ws1.column_dimensions['B'].width = 15
            
            # Sheet 2: Languages
            ws2 = wb.create_sheet("Languages")
            ws2['A1'] = 'Language'
            ws2['B1'] = 'Count'
            ws2['C1'] = 'Percentage'
            
            for cell in ['A1', 'B1', 'C1']:
                ws2[cell].fill = header_fill
                ws2[cell].font = header_font
            
            languages = analytics.get('languages', {})
            total_lang = sum(languages.values()) or 1
            
            row = 2
            for lang, count in sorted(languages.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_lang) * 100
                ws2[f'A{row}'] = lang
                ws2[f'B{row}'] = count
                ws2[f'C{row}'] = f"{percentage:.1f}%"
                row += 1
            
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 12
            ws2.column_dimensions['C'].width = 15
            
            # Sheet 3: Daily Stats
            ws3 = wb.create_sheet("Daily Usage")
            ws3['A1'] = 'Date'
            ws3['B1'] = 'Analyses'
            
            for cell in ['A1', 'B1']:
                ws3[cell].fill = header_fill
                ws3[cell].font = header_font
            
            daily = analytics.get('daily_stats', {})
            row = 2
            for date, count in sorted(daily.items(), reverse=True):
                ws3[f'A{row}'] = date
                ws3[f'B{row}'] = count
                row += 1
            
            ws3.column_dimensions['A'].width = 20
            ws3.column_dimensions['B'].width = 15
            
            # Sheet 4: Hourly Patterns
            ws4 = wb.create_sheet("Peak Hours")
            ws4['A1'] = 'Hour'
            ws4['B1'] = 'Analyses'
            
            for cell in ['A1', 'B1']:
                ws4[cell].fill = header_fill
                ws4[cell].font = header_font
            
            hourly = analytics.get('hourly_patterns', {})
            row = 2
            for hour, count in sorted(hourly.items(), key=lambda x: x[1], reverse=True):
                ws4[f'A{row}'] = hour
                ws4[f'B{row}'] = count
                row += 1
            
            ws4.column_dimensions['A'].width = 15
            ws4.column_dimensions['B'].width = 15
            
            # Save
            filename = f"MediSimplify_Analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.exports_dir, filename)
            wb.save(filepath)
            
            print(f"✅ Analytics exported: {filename}")
            return filepath
            
        except Exception as e:
            print(f"⚠️ Export error: {e}")
            return None
    
    def export_feedback(self):
        """Export feedback to Excel"""
        
        try:
            # Load feedback
            with open('user_feedback/all_feedback.json', 'r') as f:
                all_feedback = json.load(f)
            
            if not all_feedback:
                print("⚠️ No feedback to export")
                return None
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "User Feedback"
            
            # Header
            headers = ['ID', 'Date', 'Rating', 'Stars', 'Type', 'Feature', 'Language', 'Comments', 'Email']
            
            header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row, fb in enumerate(all_feedback, 2):
                ws.cell(row, 1, fb.get('id', ''))
                ws.cell(row, 2, datetime.fromisoformat(fb['timestamp']).strftime('%Y-%m-%d %H:%M'))
                ws.cell(row, 3, fb.get('rating', 0))
                ws.cell(row, 4, '★' * fb.get('rating', 0))
                ws.cell(row, 5, fb.get('feedback_type', ''))
                ws.cell(row, 6, fb.get('feature_used', ''))
                ws.cell(row, 7, fb.get('language', ''))
                ws.cell(row, 8, fb.get('comments', ''))
                ws.cell(row, 9, fb.get('email', 'anonymous'))
                
                # Color-code ratings
                rating_cell = ws.cell(row, 3)
                if fb.get('rating', 0) == 5:
                    rating_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                elif fb.get('rating', 0) <= 2:
                    rating_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            
            # Auto-width
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 50
            ws.column_dimensions['I'].width = 25
            
            # Save
            filename = f"MediSimplify_Feedback_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.exports_dir, filename)
            wb.save(filepath)
            
            print(f"✅ Feedback exported: {filename}")
            return filepath
            
        except Exception as e:
            print(f"⚠️ Export error: {e}")
            return None
    
    def export_combined(self):
        """Export both analytics and feedback in one Excel file"""
        
        try:
            # Load data
            with open('analytics/usage_stats.json', 'r') as f:
                analytics = json.load(f)
            
            with open('user_feedback/all_feedback.json', 'r') as f:
                feedback_list = json.load(f)
            
            with open('collected_data/collection_stats.json', 'r') as f:
                dataset = json.load(f)
            
            # Create workbook
            wb = Workbook()
            
            # SHEET 1: Summary
            ws1 = wb.active
            ws1.title = "Summary"
            
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=14)
            
            ws1['A1'] = 'MediSimplify - Complete Report'
            ws1['A1'].font = Font(bold=True, size=18)
            ws1.merge_cells('A1:D1')
            
            ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws1.merge_cells('A2:D2')
            
            # Usage Section
            ws1['A4'] = 'USAGE STATISTICS'
            ws1['A4'].font = header_font
            ws1['A4'].fill = header_fill
            ws1.merge_cells('A4:B4')
            
            ws1['A5'] = 'Total Analyses'
            ws1['B5'] = analytics.get('total_analyses', 0)
            ws1['A6'] = 'Prescriptions'
            ws1['B6'] = analytics.get('prescriptions_analyzed', 0)
            ws1['A7'] = 'Lab Reports'
            ws1['B7'] = analytics.get('labs_analyzed', 0)
            ws1['A8'] = 'X-Rays'
            ws1['B8'] = analytics.get('xrays_analyzed', 0)
            
            # Feedback Section
            ws1['A10'] = 'USER FEEDBACK'
            ws1['A10'].font = header_font
            ws1['A10'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            ws1.merge_cells('A10:B10')
            
            total_fb = len(feedback_list)
            ratings = [f['rating'] for f in feedback_list if f.get('rating', 0) > 0]
            avg_rating = sum(ratings) / len(ratings) if ratings else 0
            
            ws1['A11'] = 'Total Feedback'
            ws1['B11'] = total_fb
            ws1['A12'] = 'Average Rating'
            ws1['B12'] = f"{avg_rating:.2f} / 5.0"
            
            # Dataset Section
            ws1['A14'] = 'DATASET COLLECTED'
            ws1['A14'].font = header_font
            ws1['A14'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
            ws1.merge_cells('A14:B14')
            
            ws1['A15'] = 'Total Samples'
            ws1['B15'] = dataset.get('total_analyses', 0)
            ws1['A16'] = 'Prescription Images'
            ws1['B16'] = dataset.get('total_prescriptions', 0)
            ws1['A17'] = 'Lab Report Files'
            ws1['B17'] = dataset.get('total_lab_reports', 0)
            ws1['A18'] = 'X-Ray Images'
            ws1['B18'] = dataset.get('total_xrays', 0)
            
            ws1.column_dimensions['A'].width = 25
            ws1.column_dimensions['B'].width = 20
            
            # SHEET 2: Detailed Feedback
            ws2 = wb.create_sheet("All Feedback")
            
            headers = ['Date', 'Rating', 'Stars', 'Type', 'Feature', 'Language', 'Comments', 'Email']
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(1, col, header)
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            for row, fb in enumerate(feedback_list, 2):
                ws2.cell(row, 1, datetime.fromisoformat(fb['timestamp']).strftime('%Y-%m-%d %H:%M'))
                ws2.cell(row, 2, fb.get('rating', 0))
                ws2.cell(row, 3, '★' * fb.get('rating', 0))
                ws2.cell(row, 4, fb.get('feedback_type', ''))
                ws2.cell(row, 5, fb.get('feature_used', ''))
                ws2.cell(row, 6, fb.get('language', ''))
                ws2.cell(row, 7, fb.get('comments', ''))
                ws2.cell(row, 8, fb.get('email', ''))
                
                # Highlight 5-star
                if fb.get('rating', 0) == 5:
                    ws2.cell(row, 2).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            
            ws2.column_dimensions['A'].width = 18
            ws2.column_dimensions['B'].width = 8
            ws2.column_dimensions['C'].width = 10
            ws2.column_dimensions['D'].width = 18
            ws2.column_dimensions['E'].width = 15
            ws2.column_dimensions['F'].width = 12
            ws2.column_dimensions['G'].width = 60
            ws2.column_dimensions['H'].width = 25
            
            # SHEET 3: Language Distribution
            ws3 = wb.create_sheet("Languages")
            
            ws3['A1'] = 'Language'
            ws3['B1'] = 'Count'
            ws3['C1'] = 'Percentage'
            
            for cell in ['A1', 'B1', 'C1']:
                ws3[cell].fill = header_fill
                ws3[cell].font = header_font
            
            languages = analytics.get('languages', {})
            total_lang = sum(languages.values()) or 1
            
            row = 2
            for lang, count in sorted(languages.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_lang) * 100
                ws3[f'A{row}'] = lang
                ws3[f'B{row}'] = count
                ws3[f'C{row}'] = f"{percentage:.1f}%"
                row += 1
            
            ws3.column_dimensions['A'].width = 20
            ws3.column_dimensions['B'].width = 12
            ws3.column_dimensions['C'].width = 15
            
            # Save
            filename = f"MediSimplify_Complete_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.exports_dir, filename)
            wb.save(filepath)
            
            print(f"✅ Complete export: {filename}")
            return filepath
            
        except Exception as e:
            print(f"⚠️ Excel export error: {e}")
            import traceback
            traceback.print_exc()
            return None


# Initialize
excel_exporter = ExcelExporter()